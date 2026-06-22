#!/usr/bin/env python3
"""
Build the Card Lifecycle Management workbook (90-day destruction rule).

Implements the compliance spec:
  - Card_Master with calculated Card Age / Age Category / Destruction Flag
  - Dashboard with green/amber/red KPIs + aging-distribution chart + alert banner
  - Conditional formatting (green 0-59, amber 60-89, red 90+) on Card_Master & Report
  - Destruction Due Report (90+, most overdue at top)
  - Search panel (age, category, destruction status, days remaining/overdue)
  - Audit Log (populated by the VBA module in CardLifecycle_Macros.bas)
  - Config sheet (thresholds, single source of the business constants)

All aging is calculated dynamically via TODAY() - Date Received, so the file is
always current when opened. Run:  python build_workbook.py
"""

from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(HERE, "CardLifecycle.xlsx")

# ---- Business constants (mirrored on the Config sheet) ----------------------
WARNING_DAYS = 60      # 60-89  -> Warning / Nearing Expiry
ESCALATE_DAYS = 85     # 85+    -> High risk escalation
DESTRUCTION_DAYS = 90  # 90+    -> Due for Destruction
RED_ALERT_THRESHOLD = 5  # banner fires when red cards exceed this

# ---- Palette ----------------------------------------------------------------
GREEN = "C6EFCE"
GREEN_TXT = "006100"
AMBER = "FFEB9C"
AMBER_TXT = "9C6500"
RED = "FFC7CE"
RED_TXT = "9C0006"
HEADER_BG = "1F4E2C"   # WEPL-ish deep green
HEADER_TXT = "FFFFFF"
TITLE = "1F4E2C"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell):
    cell.font = Font(bold=True, color=HEADER_TXT, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def title_cell(ws, ref, text, size=16):
    ws[ref] = text
    ws[ref].font = Font(bold=True, size=size, color=TITLE)


# ============================================================================
# Sample data — varied Date Received so all three statuses are demonstrated.
# Dates are anchored relative to a reference "today" only for sane spacing;
# the workbook itself always recomputes against the real TODAY().
# ============================================================================
REF_TODAY = dt.date(2026, 6, 22)


def d(days_ago: int) -> dt.date:
    return REF_TODAY - dt.timedelta(days=days_ago)


CARD_TYPES = ["Visa Debit", "Visa Credit", "Mastercard Debit",
              "Mastercard Credit", "Prepaid"]
BRANCHES = ["Nairobi CBD", "Mombasa", "Kisumu", "Nakuru", "Eldoret", "Thika"]

# (Card Number, BIN, Card Type, Branch, days_ago_received)
SAMPLE = [
    ("4012 8888 0001", "401288", "Visa Debit", "Nairobi CBD", 5),
    ("5412 7510 0002", "541275", "Mastercard Debit", "Mombasa", 18),
    ("4012 8888 0003", "401288", "Visa Credit", "Kisumu", 33),
    ("5412 7510 0004", "541275", "Mastercard Credit", "Nakuru", 47),
    ("4012 8888 0005", "401288", "Prepaid", "Eldoret", 58),
    ("5412 7510 0006", "541275", "Visa Debit", "Thika", 61),
    ("4012 8888 0007", "401288", "Mastercard Debit", "Nairobi CBD", 67),
    ("5412 7510 0008", "541275", "Visa Credit", "Mombasa", 72),
    ("4012 8888 0009", "401288", "Prepaid", "Kisumu", 79),
    ("5412 7510 0010", "541275", "Mastercard Credit", "Nakuru", 85),
    ("4012 8888 0011", "401288", "Visa Debit", "Eldoret", 88),
    ("5412 7510 0012", "541275", "Mastercard Debit", "Thika", 89),
    ("4012 8888 0013", "401288", "Visa Credit", "Nairobi CBD", 90),
    ("5412 7510 0014", "541275", "Prepaid", "Mombasa", 95),
    ("4012 8888 0015", "401288", "Mastercard Credit", "Kisumu", 104),
    ("5412 7510 0016", "541275", "Visa Debit", "Nakuru", 118),
    ("4012 8888 0017", "401288", "Mastercard Debit", "Eldoret", 132),
    ("5412 7510 0018", "541275", "Visa Credit", "Thika", 151),
    ("4012 8888 0019", "401288", "Prepaid", "Nairobi CBD", 12),
    ("5412 7510 0020", "541275", "Mastercard Credit", "Mombasa", 26),
    ("4012 8888 0021", "401288", "Visa Debit", "Kisumu", 41),
    ("5412 7510 0022", "541275", "Mastercard Debit", "Nakuru", 63),
    ("4012 8888 0023", "401288", "Visa Credit", "Eldoret", 86),
    ("5412 7510 0024", "541275", "Prepaid", "Thika", 99),
    ("4012 8888 0025", "401288", "Mastercard Credit", "Nairobi CBD", 175),
]

# Card_Master columns
HEADERS = [
    "Card Number", "BIN", "Card Type", "Branch", "Date Received",
    "Card Age (Days)", "Age Category", "Destruction Flag",
    "Days to 90-Day Limit", "Days Overdue",
]
N = len(SAMPLE)
FIRST_DATA_ROW = 2
LAST_DATA_ROW = FIRST_DATA_ROW + N - 1


def build():
    wb = Workbook()

    # ---------------------------------------------------------------- Config
    cfg = wb.active
    cfg.title = "Config"
    title_cell(cfg, "A1", "Card Lifecycle — Configuration")
    cfg["A3"] = "Business constant"
    cfg["B3"] = "Value"
    cfg["C3"] = "Meaning"
    for c in ("A3", "B3", "C3"):
        style_header(cfg[c])
    rows = [
        ("WarningDays", WARNING_DAYS, "Age at which a card enters WARNING (nearing expiry)"),
        ("EscalateDays", ESCALATE_DAYS, "Age at which a card becomes HIGH RISK (escalate)"),
        ("DestructionDays", DESTRUCTION_DAYS, "Maximum holding period — 90+ is DUE FOR DESTRUCTION"),
        ("RedAlertThreshold", RED_ALERT_THRESHOLD, "Dashboard banner fires when red cards EXCEED this"),
    ]
    r = 4
    for name, val, meaning in rows:
        cfg[f"A{r}"] = name
        cfg[f"B{r}"] = val
        cfg[f"C{r}"] = meaning
        cfg[f"A{r}"].font = Font(bold=True)
        for col in "ABC":
            cfg[f"{col}{r}"].border = BORDER
        # named range for each constant
        wb.defined_names.add(__import__("openpyxl").workbook.defined_name.DefinedName(
            name, attr_text=f"Config!$B${r}"))
        r += 1
    cfg.column_dimensions["A"].width = 20
    cfg.column_dimensions["B"].width = 10
    cfg.column_dimensions["C"].width = 70
    cfg["A10"] = ("Card Age is computed dynamically as TODAY() - Date Received. "
                  "These constants drive every sheet; change them here only.")
    cfg["A10"].font = Font(italic=True, color="808080")

    # ------------------------------------------------------------ Card_Master
    cm = wb.create_sheet("Card_Master")
    title_cell(cm, "A1", "Card_Master")
    cm["A1"].font = Font(bold=True, size=14, color=TITLE)
    # header row at row 1? We put title via comment-free: use row 1 as headers.
    # Re-do: headers on row 1 so the Excel Table starts at row 1.
    cm["A1"] = None
    for j, h in enumerate(HEADERS, start=1):
        c = cm.cell(row=1, column=j, value=h)
        style_header(c)

    for i, (num, bin_, ctype, branch, ago) in enumerate(SAMPLE):
        row = FIRST_DATA_ROW + i
        cm.cell(row=row, column=1, value=num)
        cm.cell(row=row, column=2, value=str(bin_))
        cm.cell(row=row, column=3, value=ctype)
        cm.cell(row=row, column=4, value=branch)
        dr = cm.cell(row=row, column=5, value=d(ago))
        dr.number_format = "yyyy-mm-dd"
        # F Card Age
        cm.cell(row=row, column=6,
                value=f'=IF($E{row}="","",TODAY()-$E{row})')
        # G Age Category
        cm.cell(row=row, column=7,
                value=(f'=IF($F{row}="","",'
                       f'IF($F{row}>=DestructionDays,"Due for Destruction",'
                       f'IF($F{row}>=WarningDays,"Warning","Normal")))'))
        # H Destruction Flag
        cm.cell(row=row, column=8,
                value=f'=IF($F{row}="","",$F{row}>=DestructionDays)')
        # I Days to 90-day limit
        cm.cell(row=row, column=9,
                value=f'=IF($F{row}="","",DestructionDays-$F{row})')
        # J Days Overdue
        cm.cell(row=row, column=10,
                value=f'=IF($F{row}="","",MAX(0,$F{row}-DestructionDays))')
        for j in range(1, 11):
            cm.cell(row=row, column=j).border = BORDER
        cm.cell(row=row, column=5).number_format = "yyyy-mm-dd"

    widths = [18, 9, 18, 14, 14, 15, 20, 15, 18, 13]
    for j, w in enumerate(widths, start=1):
        cm.column_dimensions[get_column_letter(j)].width = w
    cm.freeze_panes = "A2"

    # Excel Table over the whole range (enables structured references)
    table_ref = f"A1:J{LAST_DATA_ROW}"
    tbl = Table(displayName="Card_Master", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False)
    cm.add_table(tbl)

    # Conditional formatting on the data rows (status colours). stopIfTrue so
    # the first matching rule wins. Applied to A..J so the WHOLE row colours.
    rng = f"A{FIRST_DATA_ROW}:J{LAST_DATA_ROW}"
    red_fill = PatternFill("solid", fgColor=RED)
    amber_fill = PatternFill("solid", fgColor=AMBER)
    green_fill = PatternFill("solid", fgColor=GREEN)
    cm.conditional_formatting.add(rng, FormulaRule(
        formula=[f"$F{FIRST_DATA_ROW}>=DestructionDays"], fill=red_fill,
        font=Font(color=RED_TXT), stopIfTrue=True))
    cm.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND($F{FIRST_DATA_ROW}>=WarningDays,$F{FIRST_DATA_ROW}<DestructionDays)"],
        fill=amber_fill, font=Font(color=AMBER_TXT), stopIfTrue=True))
    cm.conditional_formatting.add(rng, FormulaRule(
        formula=[f"AND($F{FIRST_DATA_ROW}<>\"\",$F{FIRST_DATA_ROW}<WarningDays)"],
        fill=green_fill, font=Font(color=GREEN_TXT), stopIfTrue=True))

    # ---------------------------------------------------------------- Helper
    # Small calc block feeding the dashboard chart (kept on Dashboard, hidden-ish).
    age_col = "Card_Master[Card Age (Days)]"

    # -------------------------------------------------------------- Dashboard
    db = wb.create_sheet("Dashboard")
    title_cell(db, "B2", "Card Lifecycle Dashboard — 90-Day Destruction Control", 18)

    # Alert banner (B4) — turns red via CF when red cards exceed threshold.
    db["B4"] = ('=IF(COUNTIF(' + age_col + ',">="&DestructionDays)>RedAlertThreshold,'
                '"⚠ ALERT: "&COUNTIF(' + age_col + ',">="&DestructionDays)&'
                '" cards are DUE FOR DESTRUCTION (threshold "&RedAlertThreshold&"). '
                'Action required.","✔ Within limits: destruction backlog under control.")')
    db.merge_cells("B4:H4")
    db["B4"].font = Font(bold=True, size=12)
    db["B4"].alignment = Alignment(horizontal="center", vertical="center")
    db.row_dimensions[4].height = 26
    db.conditional_formatting.add("B4", FormulaRule(
        formula=[f'COUNTIF({age_col},">="&DestructionDays)>RedAlertThreshold'],
        fill=PatternFill("solid", fgColor=RED), font=Font(bold=True, color=RED_TXT)))
    db.conditional_formatting.add("B4", FormulaRule(
        formula=[f'COUNTIF({age_col},">="&DestructionDays)<=RedAlertThreshold'],
        fill=PatternFill("solid", fgColor=GREEN), font=Font(bold=True, color=GREEN_TXT)))

    # KPI cards: (label, formula, fill, txt) at columns B, D, F, H on rows 6-8
    kpis = [
        ("NORMAL (0–59 days)",
         f'=COUNTIFS({age_col},">=0",{age_col},"<"&WarningDays)', GREEN, GREEN_TXT),
        ("WARNING (60–89 days)",
         f'=COUNTIFS({age_col},">="&WarningDays,{age_col},"<"&DestructionDays)', AMBER, AMBER_TXT),
        ("DUE FOR DESTRUCTION (90+)",
         f'=COUNTIF({age_col},">="&DestructionDays)', RED, RED_TXT),
        ("TOTAL CARDS",
         f'=COUNTA(Card_Master[Card Number])', "D9E1F2", "1F4E78"),
    ]
    kpi_cols = ["B", "D", "F", "H"]
    for (label, formula, fill, txt), col in zip(kpis, kpi_cols):
        nxt = chr(ord(col) + 1)
        db.merge_cells(f"{col}6:{nxt}6")
        db.merge_cells(f"{col}7:{nxt}8")
        lab = db[f"{col}6"]
        lab.value = label
        lab.font = Font(bold=True, color=txt, size=11)
        lab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lab.fill = PatternFill("solid", fgColor=fill)
        val = db[f"{col}7"]
        val.value = formula
        val.font = Font(bold=True, size=28, color=txt)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.fill = PatternFill("solid", fgColor=fill)
        for rr in (6, 7, 8):
            for cc in (col, nxt):
                db[f"{cc}{rr}"].border = BORDER
    db.row_dimensions[7].height = 36
    db.row_dimensions[8].height = 12

    # Highlighted attention lists (counts + pointer to reports)
    db["B11"] = "Cards Approaching Expiry (60–89 days)"
    db["B11"].font = Font(bold=True, color=AMBER_TXT, size=12)
    db["G11"] = f'=COUNTIFS({age_col},">="&WarningDays,{age_col},"<"&DestructionDays)'
    db["G11"].font = Font(bold=True, color=AMBER_TXT, size=12)
    db["B12"] = "Cards Due for Destruction (90+ days)"
    db["B12"].font = Font(bold=True, color=RED_TXT, size=12)
    db["G12"] = f'=COUNTIF({age_col},">="&DestructionDays)'
    db["G12"].font = Font(bold=True, color=RED_TXT, size=12)
    db["B13"] = "High-risk escalation (85+ days, not yet 90)"
    db["B13"].font = Font(bold=True, color=AMBER_TXT)
    db["G13"] = (f'=COUNTIFS({age_col},">="&EscalateDays,{age_col},"<"&DestructionDays)')
    db["B14"] = "→ See the 'Destruction_Due_Report' sheet for the actionable list."
    db["B14"].font = Font(italic=True, color="808080")

    # Chart source block (rows 17-20, cols B:C) — aging distribution
    db["B17"] = "Aging Bucket"
    db["C17"] = "Cards"
    style_header(db["B17"]); style_header(db["C17"])
    db["B18"] = "0–59 (Normal)"
    db["C18"] = f'=COUNTIFS({age_col},">=0",{age_col},"<"&WarningDays)'
    db["B19"] = "60–89 (Warning)"
    db["C19"] = f'=COUNTIFS({age_col},">="&WarningDays,{age_col},"<"&DestructionDays)'
    db["B20"] = "90+ (Destruction)"
    db["C20"] = f'=COUNTIF({age_col},">="&DestructionDays)'
    for rr in range(18, 21):
        db[f"B{rr}"].border = BORDER
        db[f"C{rr}"].border = BORDER

    chart = BarChart()
    chart.type = "col"
    chart.title = "Aging Distribution"
    chart.y_axis.title = "Number of cards"
    chart.x_axis.title = "Age bucket"
    data = Reference(db, min_col=3, min_row=17, max_row=20)
    cats = Reference(db, min_col=2, min_row=18, max_row=20)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    chart.legend = None
    db.add_chart(chart, "E17")

    db.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        db.column_dimensions[col].width = 16
    db.sheet_view.showGridLines = False

    # -------------------------------------------------- Destruction_Due_Report
    rep = wb.create_sheet("Destruction_Due_Report")
    title_cell(rep, "A1", "Destruction Due Report (90+ days) — most overdue first", 14)
    rep_headers = ["Card Number", "BIN", "Card Type", "Branch", "Date Received",
                   "Card Age", "Status", "Days Overdue"]
    for j, h in enumerate(rep_headers, start=1):
        style_header(rep.cell(row=3, column=j, value=h))
        rep.column_dimensions[get_column_letter(j)].width = 16
    # Single dynamic-array (Excel 365/2021) spill that filters + sorts.
    rep["A4"] = (
        '=IFERROR('
        'SORT('
        'FILTER('
        'HSTACK(Card_Master[Card Number],Card_Master[BIN],Card_Master[Card Type],'
        'Card_Master[Branch],Card_Master[Date Received],Card_Master[Card Age (Days)],'
        'Card_Master[Age Category],Card_Master[Days Overdue]),'
        'Card_Master[Card Age (Days)]>=DestructionDays),'
        '6,-1),'
        '"No cards are currently due for destruction.")'
    )
    rep["A4"].font = Font(color="404040")
    # Conditional formatting: all rows here are red (90+) — emphasise.
    rep.conditional_formatting.add(
        "A4:H1000",
        FormulaRule(formula=['$F4>=DestructionDays'],
                    fill=PatternFill("solid", fgColor=RED),
                    font=Font(color=RED_TXT)))
    rep["A2"] = ("Auto-generated from Card_Master. Requires Excel 365 / 2021 "
                 "(FILTER/SORT/HSTACK). For older Excel, AutoFilter Card_Master "
                 "on Age Category = 'Due for Destruction' and sort Days Overdue desc.")
    rep["A2"].font = Font(italic=True, color="808080")

    # ------------------------------------------------------------------ Search
    sr = wb.create_sheet("Search")
    title_cell(sr, "A1", "Card Search", 16)
    sr["A3"] = "Enter Card Number:"
    sr["A3"].font = Font(bold=True)
    sr["B3"] = SAMPLE[0][0]
    sr["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
    sr["B3"].border = BORDER
    sr.column_dimensions["A"].width = 26
    sr.column_dimensions["B"].width = 30

    def lk(col):
        return (f'IFERROR(XLOOKUP($B$3,Card_Master[Card Number],'
                f'Card_Master[{col}]),"")')

    fields = [
        ("Card Number", '=IFERROR(XLOOKUP($B$3,Card_Master[Card Number],Card_Master[Card Number]),"Not found")'),
        ("BIN", f'={lk("BIN")}'),
        ("Card Type", f'={lk("Card Type")}'),
        ("Branch", f'={lk("Branch")}'),
        ("Date Received", f'={lk("Date Received")}'),
        ("Card Age (Days)", f'={lk("Card Age (Days)")}'),
        ("Age Category", f'={lk("Age Category")}'),
        ("Destruction Status",
         f'=IF({lk("Card Age (Days)")}>=DestructionDays,"Yes","No")'),
        ("Days Remaining / Overdue",
         f'=IF({lk("Card Age (Days)")}="","",'
         f'IF({lk("Card Age (Days)")}>=DestructionDays,'
         f'"OVERDUE by "&({lk("Card Age (Days)")}-DestructionDays)&" days",'
         f'(DestructionDays-{lk("Card Age (Days)")})&" days remaining"))'),
    ]
    r = 5
    for label, formula in fields:
        sr[f"A{r}"] = label
        sr[f"A{r}"].font = Font(bold=True)
        sr[f"A{r}"].border = BORDER
        sr[f"B{r}"] = formula
        sr[f"B{r}"].border = BORDER
        r += 1
    sr[f"B{5+4}"].number_format = "yyyy-mm-dd"  # Date Received row
    # Colour the Age Category result row
    cat_row = 5 + 6
    sr.conditional_formatting.add(
        f"B{cat_row}",
        FormulaRule(formula=[f'$B${cat_row}="Due for Destruction"'],
                    fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TXT)))
    sr.conditional_formatting.add(
        f"B{cat_row}",
        FormulaRule(formula=[f'$B${cat_row}="Warning"'],
                    fill=PatternFill("solid", fgColor=AMBER), font=Font(color=AMBER_TXT)))
    sr.conditional_formatting.add(
        f"B{cat_row}",
        FormulaRule(formula=[f'$B${cat_row}="Normal"'],
                    fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TXT)))

    # --------------------------------------------------------------- Audit_Log
    al = wb.create_sheet("Audit_Log")
    title_cell(al, "A1", "Audit Log — destruction-eligibility status changes", 14)
    al_headers = ["Timestamp", "Card Number", "Event", "Old Status",
                  "New Status", "Age (Days)", "Performed By", "Notes"]
    for j, h in enumerate(al_headers, start=1):
        style_header(al.cell(row=3, column=j, value=h))
        al.column_dimensions[get_column_letter(j)].width = 18
    # Seed example entries documenting the three loggable transitions.
    seed = [
        (dt.datetime(2026, 4, 23, 9, 0), "5412 7510 0006", "Crossed 60 days",
         "Normal", "Warning", 60, "SYSTEM", "Early-warning flag raised"),
        (dt.datetime(2026, 5, 18, 9, 0), "5412 7510 0010", "Crossed 85 days",
         "Warning", "Warning (High Risk)", 85, "SYSTEM", "Escalation — high risk"),
        (dt.datetime(2026, 3, 24, 9, 0), "4012 8888 0013", "Crossed 90 days",
         "Warning", "Due for Destruction", 90, "SYSTEM", "Eligible for destruction"),
    ]
    r = 4
    for ts, num, ev, old, new, age, who, note in seed:
        al.cell(row=r, column=1, value=ts).number_format = "yyyy-mm-dd hh:mm"
        al.cell(row=r, column=2, value=num)
        al.cell(row=r, column=3, value=ev)
        al.cell(row=r, column=4, value=old)
        al.cell(row=r, column=5, value=new)
        al.cell(row=r, column=6, value=age)
        al.cell(row=r, column=7, value=who)
        al.cell(row=r, column=8, value=note)
        for j in range(1, 9):
            al.cell(row=r, column=j).border = BORDER
        r += 1
    al["A2"] = ("Append-only. The VBA macros in CardLifecycle_Macros.bas write here "
                "automatically when a card crosses 60 / 85 / 90 days or is marked "
                "for destruction. Do not edit historical rows (regulatory audit trail).")
    al["A2"].font = Font(italic=True, color="808080")

    # ------------------------------------------------------------------ order
    wb.move_sheet("Dashboard", -(wb.sheetnames.index("Dashboard")))  # to front
    desired = ["Dashboard", "Card_Master", "Destruction_Due_Report", "Search",
               "Audit_Log", "Config"]
    for i, name in enumerate(desired):
        wb.move_sheet(name, i - wb.sheetnames.index(name))
    wb.active = wb.sheetnames.index("Dashboard")

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({N} sample cards)")


if __name__ == "__main__":
    build()
