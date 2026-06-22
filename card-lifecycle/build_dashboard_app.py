#!/usr/bin/env python3
"""
Build an APPLICATION-GRADE dashboard for the I&M card register.

Reads the user's real register (cards_register___Copy.xlsm), imports the live
data, and produces CardOps_Dashboard.xlsx — a working "control centre" that
mirrors their 3-stage cycle:

    (1) RECORD  -> a card is logged on arrival (Card Register)
    (2) LIFECYCLE -> Days in Branch ages it: New / Normal / Warning / Due
    (3) REMIND  -> automated collection & destruction reminders fire

This .xlsx is a SAFE, self-contained preview (it does not touch the production
.xlsm, so the x14 conditional formatting, the Status dropdown, the cell-control
checkboxes and the whole VBA project on the original stay intact). The same
layout is applied in-place by CardOps_Dashboard.bas.
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.environ.get("SRC_XLSM", "/tmp/reg.xlsm")
OUT = os.path.join(HERE, "CardOps_Dashboard.xlsx")

# ---- thresholds (business rule) ---------------------------------------------
EARLY, WARN, DUE, GAP, RED_ALERT = 30, 60, 90, 7, 5

# ---- brand palette (I&M-ish deep green + gold accents) ----------------------
INK = "0E2A1E"        # near-black green (text on light)
BRAND = "1B5E20"      # deep green
BRAND_D = "0B3D17"    # darker green band
GOLD = "C8A042"       # accent
PAPER = "F4F6F4"      # page background
CARD = "FFFFFF"
MUTE = "8A938C"
GREEN = "2E7D32"; GREEN_BG = "E6F4EA"
AMBER = "F39C12"; AMBER_BG = "FEF3E0"
RED = "C0392B";   RED_BG = "FBEAE8"
BLUE = "2F6DB5";  BLUE_BG = "E8F0FA"
GREY = "5B6B61";  GREY_BG = "ECEFEC"

THINW = Side(style="thin", color="D7DDD8")
BORDER = Border(left=THINW, right=THINW, top=THINW, bottom=THINW)
NOFILL = PatternFill()


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


# =============================================================================
# 1. Load the user's real data
# =============================================================================
def load_rows():
    wb = load_workbook(SRC, data_only=True, keep_vba=True)
    ws = wb["Card Register"]
    rows = []
    for r in range(2, ws.max_row + 1):
        date = ws.cell(r, 1).value
        if date is None:
            continue
        status = (ws.cell(r, 8).value or "")
        status = str(status).strip().title() if status else ""
        rows.append({
            "date": date if isinstance(date, dt.datetime) else None,
            "name": ws.cell(r, 2).value,
            "card": ws.cell(r, 3).value,
            "phone": ws.cell(r, 5).value,
            "email": ws.cell(r, 6).value,
            "status": status,
            "issued": ws.cell(r, 9).value if isinstance(ws.cell(r, 9).value, dt.datetime) else None,
            "last_contact": ws.cell(r, 10).value if isinstance(ws.cell(r, 10).value, dt.datetime) else None,
            "notes": ws.cell(r, 11).value,
        })
    return rows


# =============================================================================
# helpers for styling
# =============================================================================
def kpi_tile(ws, anchor_col, top, label, value_formula, accent, accent_bg, icon):
    """A 3-col-wide KPI card: icon+label, big number, thin accent rule."""
    c = anchor_col
    c2 = c + 2
    L = get_column_letter
    # background card with a coloured bottom rule
    for rr in range(top, top + 4):
        for cc in range(c, c2 + 1):
            cell = ws.cell(rr, cc)
            cell.fill = fill(CARD)
            if rr == top + 3:
                cell.border = Border(bottom=Side(style="thick", color=accent))
    ws.merge_cells(f"{L(c)}{top}:{L(c2)}{top}")
    head = ws[f"{L(c)}{top}"]
    head.value = f"{icon}  {label}"
    head.font = Font(name="Segoe UI", size=10, bold=True, color=MUTE)
    head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"{L(c)}{top+1}:{L(c2)}{top+2}")
    val = ws[f"{L(c)}{top+1}"]
    val.value = value_formula
    val.font = Font(name="Segoe UI Semibold", size=30, bold=True, color=accent)
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def band(ws, rng, text, bg, txt, size=11, bold=True, align="center", h=None):
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    cell = ws[first]
    cell.value = text
    cell.fill = fill(bg)
    cell.font = Font(name="Segoe UI", size=size, bold=bold, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align == "left" else 0)
    if h:
        ws.row_dimensions[int("".join(ch for ch in first if ch.isdigit()))].height = h


# =============================================================================
# 2. Build the workbook
# =============================================================================
def build():
    rows = load_rows()
    wb = Workbook()

    # ---------- Config -------------------------------------------------------
    cfg = wb.active
    cfg.title = "Config"
    cfg["A1"] = "Card Lifecycle — Configuration"
    cfg["A1"].font = Font(bold=True, size=14, color=BRAND)
    consts = [("EarlyDays", EARLY, "New band ceiling (0-29 days)"),
              ("WarnDays", WARN, "Warning band starts (60 days)"),
              ("DueDays", DUE, "Destruction due (90 days) — the rule"),
              ("ReminderGapDays", GAP, "Min days between reminder emails (anti-spam)"),
              ("RedAlert", RED_ALERT, "Banner fires when due-for-destruction exceeds this")]
    cfg["A3"], cfg["B3"], cfg["C3"] = "Name", "Value", "Meaning"
    for col in "ABC":
        cfg[f"{col}3"].font = Font(bold=True, color="FFFFFF")
        cfg[f"{col}3"].fill = fill(BRAND)
    from openpyxl.workbook.defined_name import DefinedName
    r = 4
    for name, val, mean in consts:
        cfg[f"A{r}"], cfg[f"B{r}"], cfg[f"C{r}"] = name, val, mean
        wb.defined_names.add(DefinedName(name, attr_text=f"Config!$B${r}"))
        r += 1
    cfg.column_dimensions["A"].width = 18
    cfg.column_dimensions["B"].width = 8
    cfg.column_dimensions["C"].width = 56

    # ---------- Card Register (imported data + live lifecycle columns) -------
    reg = wb.create_sheet("Card Register")
    headers = ["Date", "Customer Name", "Card Number", "Card Type", "Phone Number",
               "Email Address", "Days in Branch", "Status", "Date Issued",
               "Last Contact", "Notes",
               "Age Band", "Lifecycle Stage", "Due for Destruction",
               "Days Overdue", "Reminders Sent", "Needs Reminder"]
    for j, h in enumerate(headers, 1):
        cell = reg.cell(1, j, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill(BRAND_D)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    reg.row_dimensions[1].height = 30

    first, last = 2, len(rows) + 1
    for i, rec in enumerate(rows):
        rr = 2 + i
        reg.cell(rr, 1, rec["date"]).number_format = "yyyy-mm-dd"
        reg.cell(rr, 2, rec["name"])
        reg.cell(rr, 3, rec["card"])
        # D Card Type via lookup (kept; will resolve once cards_type is filled)
        reg.cell(rr, 4, "")  # cards_type empty in source -> leave blank
        reg.cell(rr, 5, rec["phone"])
        reg.cell(rr, 6, rec["email"])
        # G Days in Branch — live
        reg.cell(rr, 7, f'=IF($I{rr}="",TODAY()-$A{rr},$I{rr}-$A{rr})')
        reg.cell(rr, 8, rec["status"])
        if rec["issued"]:
            reg.cell(rr, 9, rec["issued"]).number_format = "yyyy-mm-dd"
        if rec["last_contact"]:
            reg.cell(rr, 10, rec["last_contact"]).number_format = "yyyy-mm-dd"
        reg.cell(rr, 11, rec["notes"])
        # L Age Band
        reg.cell(rr, 12, (f'=IF($G{rr}>=DueDays,"Due",'
                          f'IF($G{rr}>=WarnDays,"Warning",'
                          f'IF($G{rr}>=EarlyDays,"Normal","New")))'))
        # M Lifecycle Stage
        reg.cell(rr, 13, (f'=IFS($H{rr}="Held","In Branch",'
                          f'$H{rr}="Issued","Collected",'
                          f'$H{rr}="Destroyed","Destroyed",'
                          f'TRUE,"Pending")'))
        # N Due for Destruction
        reg.cell(rr, 14, f'=AND($H{rr}="Held",$G{rr}>=DueDays)')
        # O Days Overdue
        reg.cell(rr, 15, f'=IF($N{rr},$G{rr}-DueDays,"")')
        # P Reminders Sent (from Communication Log)
        reg.cell(rr, 16, f"=COUNTIF('Communication Log'!$C:$C,$C{rr})")
        # Q Needs Reminder
        reg.cell(rr, 17, (f'=IF($H{rr}<>"Held","No",'
                          f'IF(AND($G{rr}>=EarlyDays,OR($J{rr}="",'
                          f'TODAY()-$J{rr}>=ReminderGapDays)),"Yes","No"))'))
        for j in range(1, 18):
            reg.cell(rr, j).border = BORDER

    widths = [11, 20, 16, 18, 13, 24, 9, 10, 11, 12, 22,
              9, 13, 9, 9, 8, 8]
    for j, w in enumerate(widths, 1):
        reg.column_dimensions[get_column_letter(j)].width = w
    reg.freeze_panes = "A2"
    tbl = Table(displayName="CardRegister", ref=f"A1:Q{last}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    reg.add_table(tbl)
    # status colours on the Age Band column
    bandrange = f"L{first}:L{last}"
    reg.conditional_formatting.add(bandrange, FormulaRule(formula=[f'$L{first}="Due"'], fill=fill(RED_BG), font=Font(color=RED, bold=True)))
    reg.conditional_formatting.add(bandrange, FormulaRule(formula=[f'$L{first}="Warning"'], fill=fill(AMBER_BG), font=Font(color=AMBER, bold=True)))
    reg.conditional_formatting.add(bandrange, FormulaRule(formula=[f'$L{first}="Normal"'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True)))
    reg.conditional_formatting.add(f"O{first}:O{last}", ColorScaleRule(
        start_type="num", start_value=0, start_color=AMBER_BG,
        end_type="max", end_color=RED))

    # ---------- Communication / Reminder logs (carried for live counts) ------
    com = wb.create_sheet("Communication Log")
    for j, h in enumerate(["Date Sent", "Customer Name", "Card Number",
                           "Email Address", "Email", "Type", "Sent by"], 1):
        com.cell(1, j, h).font = Font(bold=True, color="FFFFFF")
        com.cell(1, j).fill = fill(BRAND_D)
    com.sheet_properties.tabColor = GREY
    rl = wb.create_sheet("Reminder Log")
    for j, h in enumerate(["Date Sent", "Customer Name", "Email Address",
                           "Card Number", "TIB", "Status", "Email Subject"], 1):
        rl.cell(1, j, h).font = Font(bold=True, color="FFFFFF")
        rl.cell(1, j).fill = fill(BRAND_D)

    # ---------- _calc helper sheet (chart sources) ---------------------------
    calc = wb.create_sheet("_calc")
    reg_age = "CardRegister[Days in Branch]"
    reg_status = "CardRegister[Status]"
    reg_band = "CardRegister[Age Band]"
    G = f"'Card Register'!$G$2:$G${last}"
    H = f"'Card Register'!$H$2:$H${last}"
    A = f"'Card Register'!$A$2:$A${last}"

    # status mix
    calc["A1"] = "Status"; calc["B1"] = "Cards"
    statuses = ["In Branch (Held)", "Collected (Issued)", "Destroyed", "Pending"]
    skeys = ["Held", "Issued", "Destroyed", "Pend"]
    for i, (lab, key) in enumerate(zip(statuses, skeys)):
        calc.cell(2 + i, 1, lab)
        calc.cell(2 + i, 2, f'=COUNTIF({H},"{key}*")')
    # aging distribution (all cards by days in branch)
    calc["D1"] = "Age Band"; calc["E1"] = "Cards"
    bands = [("New (0-29)", f'=COUNTIFS({G},">=0",{G},"<"&EarlyDays)'),
             ("Normal (30-59)", f'=COUNTIFS({G},">="&EarlyDays,{G},"<"&WarnDays)'),
             ("Warning (60-89)", f'=COUNTIFS({G},">="&WarnDays,{G},"<"&DueDays)'),
             ("Due 90+", f'=COUNTIFS({G},">="&DueDays)')]
    for i, (lab, f_) in enumerate(bands):
        calc.cell(2 + i, 4, lab); calc.cell(2 + i, 5, f_)
    # intake by month (recording trend) for the year of the data
    yr = max((r["date"].year for r in rows if r["date"]), default=dt.date.today().year)
    calc["G1"] = "Month"; calc["H1"] = "Cards Recorded"
    for m in range(1, 13):
        lo = dt.datetime(yr, m, 1)
        hi = dt.datetime(yr + (m // 12), (m % 12) + 1, 1)
        calc.cell(1 + m, 7, lo).number_format = "mmm"
        calc.cell(1 + m, 8,
                  f'=COUNTIFS({A},">="&DATE({yr},{m},1),{A},"<"&DATE({yr+(m//12)},{(m%12)+1},1))')
    # health bar (single 100% stacked) — reuse aging bands transposed
    calc["J1"] = "seg"; calc["K1"] = "New"; calc["L1"] = "Normal"; calc["M1"] = "Warning"; calc["N1"] = "Due"
    calc["J2"] = "Portfolio"
    for i, col in enumerate(["K", "L", "M", "N"]):
        calc[f"{col}2"] = f"=_calc!$E${2+i}"
    calc.sheet_state = "hidden"

    # ---------- DASHBOARD ----------------------------------------------------
    db = wb.create_sheet("Dashboard")
    wb.move_sheet("Dashboard", -(wb.sheetnames.index("Dashboard")))
    db.sheet_view.showGridLines = False
    db.sheet_properties.tabColor = BRAND
    # page bg
    for r_ in range(1, 60):
        for c_ in range(1, 22):
            db.cell(r_, c_).fill = fill(PAPER)
    # column widths: fine grid
    for c_ in range(1, 22):
        db.column_dimensions[get_column_letter(c_)].width = 9.5
    db.column_dimensions["A"].width = 2.5
    db.column_dimensions["U"].width = 2.5

    # ----- hero band -----
    for r_ in range(2, 5):
        for c_ in range(2, 21):
            db.cell(r_, c_).fill = fill(BRAND_D)
    db.merge_cells("B2:O4")
    t = db["B2"]
    t.value = "💳  CARD OPERATIONS — CONTROL CENTRE"
    t.font = Font(name="Segoe UI Semibold", size=22, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    db.row_dimensions[2].height = 14
    db.row_dimensions[3].height = 26
    db.row_dimensions[4].height = 14
    db.merge_cells("P2:T4")
    sub = db["P2"]
    sub.value = '=TEXT(TODAY(),"dddd, dd mmm yyyy")'
    sub.font = Font(name="Segoe UI", size=11, italic=True, color=GOLD)
    sub.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    # ----- alert ribbon -----
    held_due = f'COUNTIFS({H},"Held",{G},">="&DueDays)'
    db.merge_cells("B5:T5")
    ab = db["B5"]
    ab.value = (f'=IF({held_due}>RedAlert,'
                f'"⚠  "&{held_due}&" held cards are DUE FOR DESTRUCTION (limit "&RedAlert&"). '
                f'Run the Destruction Register + final-notice reminders.",'
                f'"✓  Destruction backlog under control — "&{held_due}&" card(s) at/over 90 days.")')
    ab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    db.row_dimensions[5].height = 22
    db.conditional_formatting.add("B5", FormulaRule(formula=[f'{held_due}>RedAlert'], fill=fill(RED), font=Font(bold=True, color="FFFFFF")))
    db.conditional_formatting.add("B5", FormulaRule(formula=[f'{held_due}<=RedAlert'], fill=fill(GREEN_BG), font=Font(bold=True, color=GREEN)))

    # ----- cycle pipeline -----
    band(db, "B7:T7", "THE CARD CYCLE", PAPER, BRAND, size=11, align="left")
    db["B7"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    L = get_column_letter
    stages = [
        ("①  RECORD", "Logged on arrival", "=COUNTA(CardRegister[Card Number])", BLUE, BLUE_BG, 2),
        ("②  LIFECYCLE", "In branch, ageing", f'=COUNTIF({H},"Held*")', AMBER, AMBER_BG, 9),
        ("③  REMIND", "Reminders due now", '=COUNTIF(CardRegister[Needs Reminder],"Yes")', GREEN, GREEN_BG, 16),
    ]
    for label, desc, formula, accent, bg, c1 in stages:
        c2 = c1 + 5            # 6-col tile
        cmid = c1 + 3
        for rr in range(8, 11):
            for cc in range(c1, c2 + 1):
                db.cell(rr, cc).fill = fill(bg)
        db.cell(8, c1).border = Border(left=Side(style="thick", color=accent))
        # label
        db.merge_cells(f"{L(c1)}8:{L(cmid)}9")
        lab = db[f"{L(c1)}8"]
        lab.value = label
        lab.font = Font(name="Segoe UI Semibold", size=13, bold=True, color=accent)
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # description
        db.merge_cells(f"{L(c1)}10:{L(cmid)}10")
        d2 = db[f"{L(c1)}10"]
        d2.value = desc
        d2.font = Font(name="Segoe UI", size=9, italic=True, color=MUTE)
        d2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # big number
        db.merge_cells(f"{L(cmid+1)}8:{L(c2)}10")
        n2 = db[f"{L(cmid+1)}8"]
        n2.value = formula
        n2.font = Font(name="Segoe UI Semibold", size=24, bold=True, color=accent)
        n2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    db.row_dimensions[8].height = 18
    db.row_dimensions[9].height = 16
    db.row_dimensions[10].height = 16

    # ----- KPI tiles row -----
    band(db, "B12:T12", "KEY METRICS", PAPER, BRAND, size=11, align="left")
    db["B12"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    kpis = [
        ("TOTAL CARDS", "=COUNTA(CardRegister[Card Number])", INK, "FFFFFF", "🗂"),
        ("IN BRANCH", f'=COUNTIF({H},"Held*")', BLUE, BLUE_BG, "🏦"),
        ("COLLECTED", f'=COUNTIF({H},"Issued*")', GREEN, GREEN_BG, "✅"),
        ("DESTROYED", f'=COUNTIF({H},"Destroyed*")', GREY, GREY_BG, "🗑"),
        ("DUE FOR DESTRUCTION", f'=COUNTIFS({H},"Held",{G},">="&DueDays)', RED, RED_BG, "⛔"),
        ("NEEDS REMINDER", '=COUNTIF(CardRegister[Needs Reminder],"Yes")', AMBER, AMBER_BG, "✉"),
    ]
    col = 2
    for label, formula, accent, bg, icon in kpis:
        kpi_tile(db, col, 13, label, formula, accent, bg, icon)
        col += 3
    db.row_dimensions[13].height = 16
    db.row_dimensions[14].height = 22
    db.row_dimensions[15].height = 14
    db.row_dimensions[16].height = 10

    # ----- left: status donut -----
    band(db, "B18:I18", "PORTFOLIO STATUS", PAPER, BRAND, size=11, align="left")
    db["B18"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    donut = DoughnutChart()
    donut.title = None
    data = Reference(calc, min_col=2, min_row=1, max_row=5)
    cats = Reference(calc, min_col=1, min_row=2, max_row=5)
    donut.add_data(data, titles_from_data=True)
    donut.set_categories(cats)
    donut.holeSize = 62
    pts = [BLUE, GREEN, GREY, AMBER]
    s = donut.series[0]
    s.data_points = []
    for i, colr in enumerate(pts):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = colr
        dp.graphicalProperties.line.solidFill = "FFFFFF"
        s.data_points.append(dp)
    donut.dataLabels = DataLabelList(); donut.dataLabels.showVal = True
    donut.height, donut.width = 6.8, 11
    db.add_chart(donut, "B19")

    # ----- middle: aging distribution column -----
    band(db, "J18:R18", "AGEING DISTRIBUTION (days in branch)", PAPER, BRAND, size=11, align="left")
    db["J18"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    bar = BarChart(); bar.type = "col"; bar.title = None; bar.legend = None
    bdata = Reference(calc, min_col=5, min_row=1, max_row=5)
    bcats = Reference(calc, min_col=4, min_row=2, max_row=5)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    sb = bar.series[0]
    sb.data_points = []
    for i, colr in enumerate([GREEN, BRAND, AMBER, RED]):
        dp = DataPoint(idx=i); dp.graphicalProperties.solidFill = colr
        sb.data_points.append(dp)
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    bar.height, bar.width = 6.8, 14
    db.add_chart(bar, "J19")

    # ----- intake trend line -----
    band(db, "B33:I33", "RECORDING TREND — cards logged per month", PAPER, BRAND, size=11, align="left")
    db["B33"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    line = LineChart(); line.title = None; line.legend = None
    ldata = Reference(calc, min_col=8, min_row=1, max_row=13)
    lcats = Reference(calc, min_col=7, min_row=2, max_row=13)
    line.add_data(ldata, titles_from_data=True)
    line.set_categories(lcats)
    line.series[0].graphicalProperties.line.solidFill = BRAND
    line.series[0].graphicalProperties.line.width = 28000
    line.series[0].smooth = True
    line.height, line.width = 6.8, 14
    db.add_chart(line, "B34")

    # ----- health bar (100% stacked) -----
    band(db, "J33:R33", "BRANCH HEALTH (share by age band)", PAPER, BRAND, size=11, align="left")
    db["J33"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    hb = BarChart(); hb.type = "bar"; hb.grouping = "percentStacked"; hb.overlap = 100
    hb.title = None
    hdata = Reference(calc, min_col=11, max_col=14, min_row=1, max_row=2)
    hcats = Reference(calc, min_col=10, min_row=2, max_row=2)
    hb.add_data(hdata, titles_from_data=True)
    hb.set_categories(hcats)
    for i, colr in enumerate([GREEN, BRAND, AMBER, RED]):
        hb.series[i].graphicalProperties.solidFill = colr
    hb.height, hb.width = 3.5, 14
    hb.dataLabels = DataLabelList(); hb.dataLabels.showVal = True
    db.add_chart(hb, "J34")

    # ----- Top oldest held cards (dynamic array) -----
    band(db, "B46:T46", "OLDEST CARDS IN BRANCH — action queue (oldest first)", PAPER, BRAND, size=11, align="left")
    db["B46"].font = Font(name="Segoe UI", size=11, bold=True, color=BRAND)
    th = ["Customer", "Card Number", "Days", "Band", "Reminders", "Last Contact"]
    spans = [("B", "F"), ("G", "I"), ("J", "K"), ("L", "M"), ("N", "O"), ("P", "R")]
    for (c1, c2), h in zip(spans, th):
        db.merge_cells(f"{c1}47:{c2}47")
        cell = db[f"{c1}47"]
        cell.value = h
        cell.fill = fill(BRAND_D)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # spill the top 10 held by days desc
    db["B48"] = (
        '=IFERROR(LET(h,FILTER(CardRegister[[Customer Name]:[Reminders Sent]],'
        'CardRegister[Status]="Held"),'
        'TAKE(SORT(CHOOSECOLS(h,1,2,6,11,15),3,-1),10)),'
        '"No cards are currently Held — nothing in the action queue.")')
    db["B48"].font = Font(size=10, color=INK)

    # footer note
    db.merge_cells("B58:T58")
    fn = db["B58"]
    fn.value = ("Live workbook · ageing = TODAY()−Date Received · thresholds on Config · "
                "reminders & destruction handled by the register's VBA (SendCardEmails / "
                "GenerateDestructionRegister).")
    fn.font = Font(size=8.5, italic=True, color=MUTE)

    wb.active = wb.sheetnames.index("Dashboard")
    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(rows)} cards imported)")


if __name__ == "__main__":
    build()
